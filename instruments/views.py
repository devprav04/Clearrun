from datetime import date
from rest_framework import generics, filters, permissions
from rest_framework.exceptions import PermissionDenied
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Instrument, Vendor
from .serializers import InstrumentSerializer, VendorSerializer
from accounts.audit import log_action
from settings_app.models import CompanySettings


class VendorListCreateView(generics.ListCreateAPIView):
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer

    def perform_create(self, serializer):
        obj = serializer.save()
        log_action(self.request, 'create', 'Vendor', obj.name, f'Created vendor: {obj.name}')


class VendorDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer

    def perform_update(self, serializer):
        obj = serializer.save()
        log_action(self.request, 'update', 'Vendor', obj.name, f'Updated vendor: {obj.name}')

    def perform_destroy(self, instance):
        log_action(self.request, 'delete', 'Vendor', instance.name, f'Deleted vendor: {instance.name}')
        instance.delete()


class InstrumentListCreateView(generics.ListCreateAPIView):
    queryset = Instrument.objects.select_related('vendor').all()
    serializer_class = InstrumentSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'serial_number', 'model', 'location']
    ordering_fields = ['name', 'status', 'installation_date']

    def perform_create(self, serializer):
        obj = serializer.save()
        log_action(self.request, 'create', 'Instrument', obj.name,
                   f'Added instrument: {obj.name} ({obj.serial_number})')


class InstrumentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Instrument.objects.select_related('vendor').all()
    serializer_class = InstrumentSerializer

    def perform_update(self, serializer):
        obj = serializer.save()
        log_action(self.request, 'update', 'Instrument', obj.name,
                   f'Updated instrument: {obj.name} — status: {obj.status}')

    def perform_destroy(self, instance):
        log_action(self.request, 'delete', 'Instrument', instance.name,
                   f'Deleted instrument: {instance.name} ({instance.serial_number})')
        instance.delete()


class InstrumentNextCodeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        inst_type = request.query_params.get('type', '').upper().strip()
        if not inst_type:
            return Response({'error': 'type parameter required'}, status=400)

        s = CompanySettings.get()
        parts = [s.company_code, s.department_code, s.sub_dept_code, inst_type]
        prefix = '/'.join(p for p in parts if p)

        existing = Instrument.objects.filter(manufacturer__regex=rf'^{re.escape(prefix)}/\d+$').count()
        next_num = existing + 1
        code = f'{prefix}/{next_num}'
        return Response({'next_number': next_num, 'preview': code, 'code': code})


class InstrumentByQRView(generics.RetrieveAPIView):
    queryset = Instrument.objects.all()
    serializer_class = InstrumentSerializer
    lookup_field = 'qr_code'


INSTRUMENT_HEADERS = [
    'name', 'model', 'serial_number', 'manufacturer',
    'location', 'status', 'installation_date', 'notes',
]


class InstrumentExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Instruments'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1E3A5F')
        for col, h in enumerate(INSTRUMENT_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20

        for inst in Instrument.objects.select_related('vendor').all():
            ws.append([
                inst.name, inst.model, inst.serial_number, inst.manufacturer,
                inst.location, inst.status,
                str(inst.installation_date) if inst.installation_date else '',
                inst.notes,
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="instruments.xlsx"'},
        )


class InstrumentImportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        if request.user.role not in ('manager', 'technician'):
            raise PermissionDenied('Only managers and technicians can import instruments.')

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded.'}, status=400)

        try:
            wb = openpyxl.load_workbook(file)
        except Exception:
            return Response({'error': 'Invalid Excel file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'error': 'No data rows found.'}, status=400)

        headers = [str(h).lower().replace(' ', '_') if h else '' for h in rows[0]]
        created = updated = errors = 0
        error_list = []

        for row_idx, row in enumerate(rows[1:], 2):
            data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}
            serial = data.get('serial_number', '').strip()
            name = data.get('name', '').strip()
            if not serial or not name:
                errors += 1
                error_list.append(f'Row {row_idx}: missing name or serial_number')
                continue

            inst_data = {
                'name': name,
                'model': data.get('model', ''),
                'manufacturer': data.get('manufacturer', ''),
                'location': data.get('location', ''),
                'notes': data.get('notes', ''),
                'status': data.get('status', 'operational') or 'operational',
            }
            raw_date = data.get('installation_date', '')
            if raw_date:
                try:
                    parts = raw_date.split('-')
                    inst_data['installation_date'] = date(int(parts[0]), int(parts[1]), int(parts[2]))
                except (ValueError, IndexError):
                    pass

            try:
                obj, was_created = Instrument.objects.update_or_create(
                    serial_number=serial, defaults=inst_data
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors += 1
                error_list.append(f'Row {row_idx}: {e}')

        log_action(request, 'create', 'Instrument', 'Bulk Import',
                   f'Imported {created} new, {updated} updated, {errors} errors')
        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'error_details': error_list[:20],
        })
