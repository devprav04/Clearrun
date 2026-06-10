import io
import os
import re
import uuid
from datetime import date

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..audit import log_action
from ..deps import get_current_user, get_db
from ..models import CompanySettings, Instrument, User
from ..schemas import InstrumentCreate, InstrumentOut, InstrumentUpdate, paginated
from decouple import config

router = APIRouter(prefix='/api', tags=['instruments'])
MEDIA_ROOT = config('MEDIA_ROOT', default='media')

ALLOWED_IMPORT_EXTENSIONS = {'.xlsx', '.xls'}

STATUS_LABELS = {
    'operational': 'Operational', 'calibrating': 'Calibrating',
    'broken_down': 'Broken Down', 'scheduled_maintenance': 'Under Scheduled Maintenance',
    'out_of_service': 'Out of Service',
}

VALID_STATUSES = set(STATUS_LABELS.keys())

EXPORT_HEADERS = ['name', 'model', 'serial_number', 'manufacturer', 'location', 'status', 'installation_date', 'notes']


def _enrich(inst: Instrument) -> InstrumentOut:
    out = InstrumentOut.model_validate(inst)
    out.status_display = STATUS_LABELS.get(inst.status, inst.status)
    out.vendor_name = inst.vendor.name if inst.vendor else None
    return out


@router.get('/instruments/', response_model=dict)
def list_instruments(
    page: int = 1, page_size: int = 20,
    search: str = Query(default=''),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(Instrument)
    if search:
        like = f'%{search}%'
        q = q.filter(or_(
            Instrument.name.ilike(like),
            Instrument.serial_number.ilike(like),
            Instrument.model.ilike(like),
            Instrument.location.ilike(like),
        ))
    total = q.count()
    instruments = q.order_by(Instrument.name).offset((page - 1) * page_size).limit(page_size).all()
    return paginated([_enrich(i) for i in instruments], total)


@router.get('/instruments/next-code/')
def next_instrument_code(inst_type: str = Query(...), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    s = db.query(CompanySettings).filter(CompanySettings.id == 1).first()
    if s:
        parts = [s.company_code, s.department_code, s.sub_dept_code, inst_type.upper()]
    else:
        parts = [inst_type.upper()]
    prefix = '/'.join(p for p in parts if p)
    # Query qr_code (where the generated code is stored), not manufacturer
    pattern = rf'^{re.escape(prefix)}/\d+$'
    existing = db.query(Instrument).filter(Instrument.qr_code.op('~')(pattern)).count()
    next_num = existing + 1
    code = f'{prefix}/{next_num}'
    return {'next_number': next_num, 'preview': code, 'code': code}


@router.get('/instruments/export/')
def export_instruments(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Instruments'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    for col, h in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20
    for inst in db.query(Instrument).order_by(Instrument.name).all():
        ws.append([inst.name, inst.model, inst.serial_number, inst.manufacturer,
                   inst.location, inst.status,
                   str(inst.installation_date) if inst.installation_date else '',
                   inst.notes])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="instruments.xlsx"'},
    )


@router.post('/instruments/import/')
def import_instruments(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ('manager', 'technician') and not current_user.is_superuser:
        raise HTTPException(status_code=403, detail='Only managers and technicians can import instruments.')

    # Validate file type by extension and content-type
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        raise HTTPException(status_code=400, detail='Only .xlsx and .xls files are accepted.')

    try:
        wb = openpyxl.load_workbook(file.file)
    except Exception:
        raise HTTPException(status_code=400, detail='Invalid or corrupt Excel file.')
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail='No data rows found in the file.')

    headers = [str(h).lower().replace(' ', '_') if h else '' for h in rows[0]]
    created = updated = errors = 0
    error_list = []

    for row_idx, row in enumerate(rows[1:], 2):
        data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers)}
        serial = data.get('serial_number', '').strip()
        name   = data.get('name', '').strip()
        if not serial or not name:
            errors += 1
            error_list.append(f'Row {row_idx}: missing name or serial_number')
            continue

        status = data.get('status', 'operational').strip() or 'operational'
        if status not in VALID_STATUSES:
            status = 'operational'

        inst_data = {
            'name': name,
            'model': data.get('model', ''),
            'manufacturer': data.get('manufacturer', ''),
            'location': data.get('location', ''),
            'notes': data.get('notes', ''),
            'status': status,
        }
        raw_date = data.get('installation_date', '').strip()
        if raw_date:
            try:
                parts = raw_date.split('-')
                inst_data['installation_date'] = date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                pass  # leave installation_date unset; don't fail the row

        try:
            existing = db.query(Instrument).filter(Instrument.serial_number == serial).first()
            if existing:
                for k, v in inst_data.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                inst_data['serial_number'] = serial
                inst_data.setdefault('qr_code', f'CR-{serial}')
                db.add(Instrument(**inst_data))
                created += 1
            db.commit()
        except Exception as e:
            db.rollback()
            errors += 1
            error_list.append(f'Row {row_idx}: {e}')

    log_action(db, current_user, request, 'create', 'Instrument', 'Bulk Import',
               f'Imported {created} new, {updated} updated, {errors} errors')
    return {'created': created, 'updated': updated, 'errors': errors, 'error_details': error_list[:20]}


@router.get('/instruments/qr/{qr_code}/', response_model=InstrumentOut)
def get_by_qr(qr_code: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    inst = db.query(Instrument).filter(Instrument.qr_code == qr_code).first()
    if not inst:
        raise HTTPException(status_code=404, detail='Instrument not found.')
    return _enrich(inst)


@router.get('/instruments/{instrument_id}/', response_model=InstrumentOut)
def get_instrument(instrument_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(status_code=404, detail='Instrument not found.')
    return _enrich(inst)


@router.post('/instruments/', response_model=InstrumentOut, status_code=201)
def create_instrument(
    request: Request, body: InstrumentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = body.model_dump()
    serial = data['serial_number']
    if db.query(Instrument).filter(Instrument.serial_number == serial).first():
        raise HTTPException(status_code=400, detail=f'Serial number {serial!r} already exists.')
    data.setdefault('qr_code', f'CR-{serial}')
    inst = Instrument(**data)
    db.add(inst)
    db.commit()
    db.refresh(inst)
    log_action(db, current_user, request, 'create', 'Instrument', inst.name,
               f'Added instrument: {inst.name} ({inst.serial_number})')
    return _enrich(inst)


@router.patch('/instruments/{instrument_id}/', response_model=InstrumentOut)
def update_instrument(
    instrument_id: int, request: Request, body: InstrumentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(status_code=404, detail='Instrument not found.')
    updates = body.model_dump(exclude_none=True)
    # Guard: don't let another instrument steal a serial_number
    if 'serial_number' in updates:
        collision = db.query(Instrument).filter(
            Instrument.serial_number == updates['serial_number'],
            Instrument.id != instrument_id,
        ).first()
        if collision:
            raise HTTPException(status_code=400, detail=f'Serial number {updates["serial_number"]!r} already in use.')
    for field, value in updates.items():
        setattr(inst, field, value)
    db.commit()
    db.refresh(inst)
    log_action(db, current_user, request, 'update', 'Instrument', inst.name,
               f'Updated instrument: {inst.name} — status: {inst.status}')
    return _enrich(inst)


@router.delete('/instruments/{instrument_id}/', status_code=204)
def delete_instrument(
    instrument_id: int, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(status_code=404, detail='Instrument not found.')
    log_action(db, current_user, request, 'delete', 'Instrument', inst.name,
               f'Deleted instrument: {inst.name} ({inst.serial_number})')
    db.delete(inst)
    db.commit()
